import numpy as np
import pandas as pd
import openpyxl
import os
import sys
from os import listdir
from os.path import isfile, isdir, join
from win32com.client import Dispatch
import time
from datetime import datetime, timedelta
import requests
import json
import holidays

database_path = join(os.path.dirname(os.getcwd()), 'database')

def reload():

    global database_path
    folder_names = [folder
                    for folder in listdir(database_path)
                    if isdir(join(database_path, folder))]

    for folder in folder_names:
        file_names = [file
                      for file in listdir(join(database_path, folder))
                      if isfile(join(database_path, folder, file))]

        for file in file_names:
            excel = Dispatch("Excel.Application")
            excel.Visible = True
            for wb in [wb for wb in excel.Workbooks]:
                wb.Close(True)
            excel.Workbooks.Open(os.path.join(database_path, file))
            time.sleep(3) # suspend 3 secs for excel to catch up python
            excel.Range("A1:XFD1048576").Select()
            excel.Selection.Copy()
            excel.Selection.PasteSpecial(Paste=-4163)
            excel.ActiveWorkbook.Save()
            excel.ActiveWorkbook.Close()

def request_period_fs(year=int, quarter=int, fs_type=str):

    global database_path
    folder = 'fs_general_industry'
    file = fs_type + '_' + str(year) + 'q' + str(quarter) + '.xlsm'

    # create Workbook object, select active Worksheet
    raw_fiinpro \
        = openpyxl.load_workbook(
        os.path.join(database_path, folder, file)).active

    try:
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row-21,
                                amount=1000)
    except IndexError:
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        for wb in [wb for wb in excel.Workbooks]:
            wb.Close(True)
        excel.Workbooks.Open(os.path.join(database_path, file))
        time.sleep(3) # suspend 3 secs for excel to catch up python
        excel.Range("A1:XFD1048576").Select()
        excel.Selection.Copy()
        excel.Selection.PasteSpecial(Paste=-4163)
        excel.ActiveWorkbook.Save()
        excel.ActiveWorkbook.Close()
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21,
                                amount=1000)

    # get file info
    date_of_extract = raw_fiinpro['B6'].value
    report_year = raw_fiinpro['E8'].value[-25:-21]
    report_quarter = raw_fiinpro['E8'].value[-11:-10]
    if file[2] == '_':
        fs_type = file[:2]
    else:
        fs_type = file[:3]
    # delete header rows
    raw_fiinpro.delete_rows(idx=0, amount=7)
    raw_fiinpro.delete_rows(idx=2, amount=1)

    # import to DataFrame, no column labels, no index
    clean_data = pd.DataFrame(raw_fiinpro.values)

    # convert Fiinpro names to plain_text
    for i in range(len(clean_data.iloc[0])):
        if len(clean_data.iloc[0,i]) > 45:
            clean_data.iloc[0,i] = \
                clean_data.iloc[0,i][:-45].split('.')[-1]\
                    .lstrip(' ').rstrip(' ').lower()\
                    .replace('-','_').replace(' ','_')\
                    .replace('___','_').replace('__','_')\
                    .replace('in_which:_','')
        else:
            pass

    # assign column labels and index
    clean_data.columns = clean_data.iloc[0,:]
    clean_data.drop(index=[0], inplace=True)
    clean_data.index \
        = pd.MultiIndex.from_arrays(
        [[report_year]*len(clean_data),
        [report_quarter]*len(clean_data),
        clean_data['Ticker'].tolist()]
    )
    clean_data.index.set_names(['year', 'quarter', 'ticker'], inplace=True)

    # rename 2 columns
    clean_data.rename(
        columns={'Name':'full_name', 'Exchange':'exchange'},
        inplace=True
    )

    # drop unwanted columns and index
    clean_data.drop(
        columns=['No', 'Ticker'],
        inplace=True
    )

    # fill na with 0s
    clean_data.fillna(0, inplace=True)

    if fs_type == 'bs':
        # handle duplicated "cost" (BS)
        for i in range(len(clean_data.columns)):
            if clean_data.columns.values[i] == 'cost':
                clean_data.columns.values[i] \
                    = str(clean_data.columns.values[i-1]) + '_cost'
            else:
                pass

        # handle duplicated "accumulated_depreciation" (BS)
        for i in range(len(clean_data.columns)):
            if clean_data.columns.values[i] == 'accumulated_depreciation':
                clean_data.columns.values[i] = \
                    str(clean_data.columns.values[i-2]) \
                    + '_accumulated_depreciation'
            else:
                pass

        # handle duplicated "held_to_maturity_investment" (BS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='held_to_maturity_investment')
            | (clean_data.columns.duplicated()),
            'short_term_held_to_maturity_investment'
        )

        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='held_to_maturity_investment'),
            'long_term_held_to_maturity_investment'
        )

        # handle duplicated "other_current_assets" (BS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='other_current_assets')
            | (clean_data.columns.duplicated()),
            'total_other_current_assets'
        )

        # handle duplicated "long_term_trade_receivables" (BS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='long_term_trade_receivables')
            | (clean_data.columns.duplicated()),
            'total_long_term_trade_receivables'
        )

        # handle duplicated "other_long_term_assets" (BS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='other_long_term_assets')
            | (clean_data.columns.duplicated()),
            'total_other_long_term_assets'
        )

        # handle duplicated "government_bonds_purchased_for_resale" (BS)
        clean_data.columns\
            = clean_data.columns.where(
            ~(clean_data.columns=='government_bonds_purchased_for_resale')
            | (clean_data.columns.duplicated()),
            'government_bonds_purchased_for_resale_(asset)'
        )

        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='held_to_maturity_investment'),
            'long_term_held_to_maturity_investment_(liability)'
        )

        # handle duplicated "preferred_shares" (BS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='preferred_shares')
            | (clean_data.columns.duplicated()),
            'preferred_shares_(libability)'
        )

        # handle duplicated "preferred_shares" (IS)
        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='preferred_shares')
            | (clean_data.columns.duplicated()),
            'preferred_shares_(libability)'
        )

        clean_data.columns \
            = clean_data.columns.where(
            ~(clean_data.columns=='preferred_shares'),
            'preferred_shares_(equity)'
        )

        # handle 1_common_share
        clean_data.rename(
            columns={'1_common_shares': 'common_shares'},
            inplace=True
        )

    else:
        pass

    clean_data.columns \
        = pd.MultiIndex.from_product(
        [[fs_type], clean_data.columns.tolist()]
    )

    print('Data Extracted!')
    return clean_data


def request_fs(): # this function need improving

    global database_path
    folder = 'fs_general_industry'
    fs_types = list(dict.fromkeys([name.split('_')[0]
                                   for name in listdir(join(database_path,
                                                           folder))
                                   if isfile(join(database_path,
                                                  folder,
                                                  name))]))
    for fs_type in fs_types:
        if fs_type.startswith('~$'):
            fs_types.remove(fs_type)

    periods = request_period_list()

    frames = list()
    for period in periods:
        for fs_type in fs_types:
            try:
                frames.append(
                    request_period_fs(int(period[:4]),
                                      int(period[-1]),
                                      fs_type)
                )
            except FileNotFoundError:
                continue

    df = pd.concat(frames, axis=1, join='outer')\
        .groupby(axis=1, level=[1], dropna=False, sort=False)\
        .sum(min_count=1)
    del df['full_name']
    del df['exchange']

    return df


def request_ticker_fs(ticker=str):

    global database_path
    folder = 'fs_general_industry'

    file_names = [
        f for f in listdir(join(database_path, folder))
        if isfile(join(database_path, folder, f))
    ]

    refs = [
        (int(name[-11:-7]),int(name[-6]),name[:2]
        if name[2]=='_' else name[:3])
        for name in file_names
    ]

    fs_types \
        = list(dict.fromkeys([name.split('_')[0]
                              for name in listdir(join(database_path,
                                                       folder))
                              if isfile(join(database_path, folder, name))]))
    fs_types.sort()

    for fs_type in fs_types:
        if fs_type.startswith('~$'):
            fs_types.remove(fs_type)

    inds = list()
    for fs_type in fs_types:
        inds.append(
            request_period_fs(year=refs[-1][0],
                              quarter=refs[-1][1],
                              fs_type=fs_type) \
                .xs(ticker, axis=0, level=2) \
                .drop(['full_name', 'exchange'], level=1, axis=1).T \
                .index.tolist())

    inds = dict([(fs_type, ind) for fs_type, ind in zip(fs_types, inds)])

    fs = pd.concat(
        [
            request_period_fs(year=ref[0], quarter=ref[1], fs_type=ref[2])
                .xs(ticker, axis=0, level=2)
                .drop(['full_name', 'exchange'], axis=1).T
                .set_index(pd.MultiIndex.from_product(
                    [[ref[2]], inds[ref[2]]]))
            for ref in refs
        ]
    )
    fs = fs.groupby(fs.index, sort=False).sum()

    print('Data Extracted!')
    return fs


def request_ticker_list():

    last_period = request_period_list()[-1]
    tickerlist \
        = request_period_fs(year=int(last_period[:4]),
                            quarter=int(last_period[-1]),
                            fs_type='is')\
        .index.get_level_values(level=2).tolist()

    return tickerlist


def request_variable_names(fs_type=str):

    last_period = request_period_list()[-1]
    variable_names \
        = request_period_fs(year=int(last_period[:4]),
                            quarter=int(last_period[-1]),
                            fs_type=fs_type)\
        .columns.get_level_values(level=1).tolist()
    try:
        variable_names.remove('full_name')
        variable_names.remove('exchange')
    except ValueError:
        pass

    return variable_names


def request_period_list():

    folder = 'fs_general_industry'
    periods \
        = list(dict.fromkeys(
        [name[-11:-5] for name in listdir(join(database_path, folder))
         if isfile(join(database_path, folder, name))]))
    periods.sort()

    return periods


def request_industry(standard=str):

    global database_path
    standards = request_industry_standard()
    st_dict = dict()
    folder = 'industry_classification'

    for st in standards:
        # create Workbook object, select active Worksheet
        raw_bloomberg \
            = openpyxl.load_workbook(
            os.path.join(database_path, folder, st+'.xlsx')
        ).active
        # delete Bloomberg Sign
        raw_bloomberg.delete_rows(idx=raw_bloomberg.max_row)
        # delete headers
        raw_bloomberg.delete_rows(idx=0, amount=2)
        clean_data = pd.DataFrame(raw_bloomberg.values)
        clean_data.iloc[1] = clean_data.iloc[0]
        clean_data.drop(index=0, inplace=True)
        # set index and columns
        clean_data.index = clean_data.iloc[:,0]
        clean_data.index.rename('ticker')
        clean_data.columns = clean_data.iloc[0,:]
        clean_data.columns.rename('level')
        # remore unwanted columns, rows
        clean_data.drop(columns=['Ticker', 'Name'],
                        index=['Ticker'], inplace=True)
        # rename columns
        clean_data.columns\
            = pd.Index(
            data = [
                clean_data.columns[i].split()[0].lower()
                    .split(' ', maxsplit=1)[0] + '_l' + str(i+1)
                for i in range(clean_data.shape[1])
            ]
        )
        # rename index
        clean_data.index \
            = pd.Index(
            data=[clean_data.index[i].split()[0]
                for i in range(clean_data.shape[0])]
        )

        st_dict[st] = clean_data

    return st_dict[standard]


def request_industry_standard():
    global database_path
    folder = 'industry_classification'
    standards \
        = list(dict.fromkeys(
        [name[:-5] for name in listdir(join(database_path, folder))
         if isfile(join(database_path, folder, name))]))
    return standards


def request_industry_level(standard=str):
    levels = request_industry(standard).columns.tolist()
    return levels


def request_industry_list(standard=str, level=int):
    industries \
        = request_industry(standard)[standard + '_l' + str(level)]\
        .drop_duplicates().tolist()
    return industries


def request_financial_ticker(): # could be improved
    l1 = list()
    for standard in request_industry_standard():
        l1.append(request_industry_list(standard,1))
    financial_names = list()
    for list_ in l1:
        for industry in list_:
            if 'inancial' in industry:
                financial_names.append(industry)
            else:
                continue
    financial_tickers_ = list()
    for standard, fin in zip(request_industry_standard(),financial_names):
        full_list = request_industry(standard)
        financial_tickers_\
            .append(
            full_list.loc[full_list[standard+'_l'+str(1)] == fin]
                .index.tolist()
        )
    financial_tickers = list()
    for i in range(len(financial_tickers_)):
        for j in range(len(financial_tickers_[i])):
            financial_tickers.append(financial_tickers_[i][j])

    financial_tickers = list(dict.fromkeys(financial_tickers))
    financial_tickers.sort()

    return financial_tickers


def request_return():

    periods = request_period_list()
    tickers = request_ticker_list()

    returns = pd.DataFrame(data=np.zeros((len(tickers),len(periods))),
                           columns=[periods[i]
                                    for i in range(len(periods))],
                           index=[tickers[j]
                                  for j in range(len(tickers))])

    date_input = {'q1':['01-01','03-31'], 'q2':['04-01','06-30'],
                  'q3':['07-01','09-30'], 'q4':['10-01','12-28']}
    # Note: because of the 'Bday' function structure,
    #       December must be adjusted to 28 instead of 31

    price_data = dict()
    for ticker in tickers:
        try:
            price_data[ticker] = request_trading_hist(ticker)
        except (KeyError, IndexError):
            continue

    def Bday(date=str):
        date_ = datetime(year=int(date.split('-')[0]),
                         month=int(date.split('-')[1]),
                         day=int(date.split('-')[2]))
        one_day = timedelta(days=1)
        while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
            date_ = date_ + one_day

        return date_.strftime(format='%Y-%m-%d')

    for period in periods:
        for ticker in tickers:
            try:
                fromdate = Bday(period[:4] + '-' + date_input[period[-2:]][0])
                todate = Bday(period[:4] + '-' + date_input[period[-2:]][1])
            except (KeyError, IndexError):
                continue
            try:
                f = price_data[ticker]['trading_date'] == fromdate
                t = price_data[ticker]['trading_date'] == todate
                try:
                    open_price = price_data[ticker].loc[f]['close'].iloc[0]
                    close_price = price_data[ticker].loc[t]['close'].iloc[0]
                    returns.loc[ticker, period] = close_price / open_price - 1
                except (IndexError, KeyError, RuntimeWarning):
                    returns.loc[ticker, period] = np.nan
            except (IndexError, KeyError, RuntimeWarning):
                returns.loc[ticker, period] = np.nan

    returns.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)

    return returns


def request_price():

    periods = request_period_list()
    tickers = request_ticker_list()

    prices = pd.DataFrame(data=np.zeros((len(tickers),len(periods))),
                          columns=[periods[i]
                                   for i in range(len(periods))],
                          index=[tickers[j]
                                 for j in range(len(tickers))])

    date_input = {'q1':'03-31', 'q2':'06-30',
                  'q3':'09-30', 'q4':'12-28'}
    # Note: because of the 'Bday' function structure,
    #       December must be adjusted to 28 instead of 31

    price_data = dict()
    for ticker in tickers:
        try:
            price_data[ticker] = request_trading_hist(ticker)
        except (KeyError, IndexError):
            continue

    def Bday(date=str):
        date_ = datetime(year=int(date.split('-')[0]),
                         month=int(date.split('-')[1]),
                         day=int(date.split('-')[2]))
        one_day = timedelta(days=1)
        while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
            date_ = date_ + one_day

        return date_.strftime(format='%Y-%m-%d')

    for period in periods:
        for ticker in tickers:
            try:
                date = Bday(period[:4] + '-' + date_input[period[-2:]])
            except (KeyError, IndexError):
                continue
            try:
                t = price_data[ticker]['trading_date'] == date
                try:
                    prices.loc[ticker, period] \
                        = price_data[ticker].loc[t]['close'].iloc[0]
                except (IndexError, KeyError, RuntimeWarning):
                    prices.loc[ticker, period] = np.nan
            except (IndexError, KeyError, RuntimeWarning):
                prices.loc[ticker, period] = np.nan

    prices.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)
    return prices


def ownership_structure():

    global database_path
    folder = 'ownership'
    file = [f for f in listdir(join(database_path, folder))
            if isfile(join(database_path, folder, f))][-1]

    excel = Dispatch("Excel.Application")
    excel.Visible = True
    for wb in [wb for wb in excel.Workbooks]:
        wb.Close(True)
    excel.Workbooks.Open(os.path.join(database_path, file))
    time.sleep(3) # suspend 3 secs for excel to catch up python
    excel.Range("A1:XFD1048576").Select()
    excel.Selection.Copy()
    excel.Selection.PasteSpecial(Paste=-4163)
    excel.ActiveWorkbook.Save()
    excel.ActiveWorkbook.Close()

    # create Workbook object, select active Worksheet
    raw_fiinpro = openpyxl.load_workbook(
        os.path.join(database_path, folder, file)
    ).active
    try:
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row-21, amount=1000)
    except IndexError:
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        for wb in [wb for wb in excel.Workbooks]:
            wb.Close(True)
        excel.Workbooks.Open(os.path.join(database_path, file))
        time.sleep(3) # suspend 3 secs for excel to catch up python
        excel.Range("A1:XFD1048576").Select()
        excel.Selection.Copy()
        excel.Selection.PasteSpecial(Paste=-4163)
        excel.ActiveWorkbook.Save()
        excel.ActiveWorkbook.Close()
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21,
                                amount=1000)

    # get file info
    date_of_extract = raw_fiinpro['B6'].value

    # delete header rows
    raw_fiinpro.delete_rows(idx=0, amount=7)
    raw_fiinpro.delete_rows(idx=2, amount=1)

    # import to DataFrame
    clean_data = pd.DataFrame(raw_fiinpro.values)

    # drop unwanted index and columns
    clean_data.drop(columns=[0], inplace=True)
    clean_data.drop(index=[0], inplace=True)

    # set ticker as index
    clean_data.index = clean_data.iloc[:,0]
    clean_data.drop(clean_data.columns[0], axis=1, inplace=True)

    # rename columns
    columns = ['full_name', 'exchange', 'state_share',
               'state_percent', 'frgn_share', 'frgn_percent',
               'other_share', 'other_percent', 'frgn_maxpercent',
               'frgn_maxshare','frgn_remainshare']
    clean_data.columns = columns

    return clean_data, date_of_extract


def request_trading_hist(ticker=str, fromdate=None, todate=None):
    address = 'https://api.phs.vn/market/utilities.svc/GetShareIntraday'
    pd.options.mode.chained_assignment = None
    if fromdate is not None and todate is not None:
        if datetime.strptime(fromdate, '%Y-%m-%d') < \
                datetime(year=2015, month=1, day=1):
            raise Exception('Only data since 2015/01/01 is reliable')
        else:
            try:
                r = requests.post(address,
                    data=json.dumps(
                        {'symbol': ticker,
                         'fromdate': fromdate,
                         'todate': todate}
                    ),
                    headers={'content-type': 'application/json'})
                history = pd.DataFrame(json.loads(r.json()['d']))
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
    else:
        try:
            r = requests.post(address,
                              data=json.dumps(
                                  {'symbol': ticker,
                                   'fromdate': datetime(year=2015,
                                                        month=1,
                                                        day=1) \
                                       .strftime("%Y-%m-%d"),
                                   'todate': datetime.now()
                                       .strftime("%Y-%m-%d")}
                              ),
            headers={'content-type': 'application/json'})
            history = pd.DataFrame(json.loads(r.json()['d']))
        except KeyError:
            try:
                r = requests.post(address,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': (datetime.now()
                                                    - timedelta(days=1000)) \
                                           .strftime("%Y-%m-%d"),
                                       'todate': datetime.now()
                                           .strftime("%Y-%m-%d")}
                                  ),
                                  headers={'content-type': 'application/json'})
                history = pd.DataFrame(json.loads(r.json()['d']))
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

    history.rename(columns={'symbol':'ticker',
                            'prior_price':'ref',
                            'close_price':'close'}, inplace=True)

    def addzero(int_=str):
        if len(int_) == 1:
            int_ = '0' + int_
        else:
            pass
        return int_

    for i in range(history.shape[0]):
        history['trading_date'].iloc[i] \
            = history['trading_date'].iloc[i][:-12].split('/')[2] + "-"\
              + addzero(history['trading_date'].iloc[i][:-12] \
                        .split('/')[0]) + "-" \
              + addzero(history['trading_date'].iloc[i][:-12] \
                        .split('/')[1])

    history['ref'].iloc[0] = history['close'].iloc[0] # wait for 'open' from IT

    for col in ['ref', 'close', 'high', 'low']: # wait 'open' from IT
        for i in range(1, history.shape[0]):
            if history[col].iloc[i] == 0:
                history[col].iloc[i] = history[col].iloc[i-1]

    return history


def request_trading_intra(ticker=str, fromdate=None, todate=None):
    address = 'https://api.phs.vn/market/Utilities.svc/GetRealShareIntraday'
    pd.options.mode.chained_assignment = None
    if fromdate is not None and todate is not None:
        if datetime.strptime(todate, '%Y-%m-%d') \
                - datetime.strptime(fromdate, '%Y-%m-%d') > timedelta(days=60):
            raise Exception('Can\'t extract more than 60 days')
        else:
            try:
                r = requests.post(address,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': fromdate,
                                       'todate': todate}
                                  ),
                                  headers={'content-type': 'application/json'})
                #history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
                intraday = pd.DataFrame(json.loads(r.json()['d'])['intradays'])
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
    else:
        try:
            fromdate = (datetime.now()
                        - timedelta(days=60)).strftime("%Y-%m-%d")
            todate = (datetime.now()
                      - timedelta(days=0)).strftime("%Y-%m-%d")
            r = requests.post(address,
                              data=json.dumps(
                                  {'symbol': ticker,
                                   'fromdate': fromdate,
                                   'todate': todate}
                              ),
                              headers={'content-type': 'application/json'})
            #history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
            intraday = pd.DataFrame(json.loads(r.json()['d'])['intradays'])
        except KeyError:
            raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

    def datemod(date=str):
        def addzero(int_=str):
            if len(int_) == 1:
                int_ = '0' + int_
            else:
                pass
            return int_
        day = addzero(date.split('/')[0])
        month = addzero(date.split('/')[1])
        date = date.split('/')[2][:4] + '-' \
               + month + '-' \
               + day + ' ' \
               + date.split(" ",maxsplit=1)[1]
        return date

    for i in range(intraday.shape[0]):
        intraday['trading_time'].iloc[i] \
            = datemod(intraday['trading_time'].iloc[i])

    return intraday

