import numpy as np
import pandas as pd
import openpyxl
import os
import sys
from os import listdir
from os.path import isfile, isdir, join
from win32com.client import Dispatch
import time
from datetime import datetime, timedelta
import requests
import json
import holidays
from typing import Union

database_path \
    = join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))),
           'database')


def request_fstype() -> list:

    """
    This function returns valid financital statements

    :param: None
    :return: list
    """

    folders = [f for f in listdir(database_path) if f.startswith('fs_')]
    fs_types = []
    for folder in folders:
        fs_types = [name.split('_')[0]
                    for name in listdir(join(database_path, folder))
                    if isfile(join(database_path, folder, name))]
        fs_types = list(dict.fromkeys(fs_types))
    fs_types = [x for x in fs_types if not x.startswith('~$')]
    fs_types.sort()

    return fs_types


def request_segment_all() -> list:

    """
    This function returns the names of segments

    :param: None
    :return: list
    """

    folders = [f for f in listdir(database_path) if f.startswith('fs_')]
    segments = [x.split('_')[1] for x in folders]
    segments.sort()

    return segments


def request_exchange_all() -> pd.DataFrame:

    """
    This function returns stock exchanges of all tickers

    :param: None
    :return: pandas.DataFrame
    """

    periods = request_period()
    latest_period = periods[-1]
    a = pd.DataFrame(columns=['exchange'])
    for segment in request_segment_all():
        table = request_fs_period(int(latest_period[:4]),
                                  int(latest_period[-1]),
                                  segment, 'is', 'all')
        table = table.xs(key='exchange', axis=1, level=1)
        table = table.droplevel(level=['year', 'quarter'],)
        table.columns = ['exchange']
        a = pd.concat([a,table])

    return a


def request_exchange(ticker) -> str:

    """
    This function returns stock exchange of given stock

    :param ticker: stock's ticker
    :type ticker: str
    :return: str
    """

    exchange_table = request_exchange_all()
    exchange = exchange_table.loc[ticker].iloc[0]
    return exchange


def request_financial_ticker(sector_break=False) \
        -> Union[list, dict]:

    """
    This function returns all tickers of financial segments

    :param sector_break: False: ignore sectors, True: show sectors
    :return: list (sector_break=False), dictionary (sector_break=True)
    """

    financials = ['bank', 'sec', 'ins']
    periods = request_period()
    latest_period = periods[-1]

    tickers = []
    tickers_ = dict()
    for segment in financials:
        folder = 'fs_' + segment + '_industry'
        file = 'is_' + latest_period[:4] + 'q' + latest_period[-1] + '.xlsm'
        raw_fiinpro \
            = openpyxl.load_workbook(
            os.path.join(database_path, folder, file)).active
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row-21,
                                amount=1000)
        # delete headers
        raw_fiinpro.delete_rows(idx=0, amount=7)
        raw_fiinpro.delete_rows(idx=2, amount=1)
        # import
        clean_data = pd.DataFrame(raw_fiinpro.values)
        clean_data.drop(index=[0], inplace=True)
        # remove OTC
        a = clean_data.loc[:, 3] != 'OTC'
        if sector_break is False:
            tickers += clean_data.loc[:,1][a].tolist()
        else:
            tickers_[segment] = clean_data.loc[:,1][a].tolist()
            tickers = tickers_

    return tickers


def request_segment(ticker=str) -> str:

    """
    This function returns the segment of a given ticker

    :param ticker: stock's ticker
    :return: str
    """

    segment = ''
    financial_tickers = request_financial_ticker()
    if ticker not in financial_tickers:
        segment = 'gen'
    else:
        financial_tickers = request_financial_ticker(True)
        for key in financial_tickers.keys():
            if ticker not in financial_tickers[key]:
                pass
            else:
                segment = key
                break

    return segment


def reload() -> None:

    """
    This function handles cached data in newly-added files

    :param: None
    :return: None
    """

    global database_path
    folder_names = [folder
                    for folder in listdir(database_path)
                    if isdir(join(database_path, folder))]

    for folder in folder_names:
        file_names = [file
                      for file in listdir(join(database_path, folder))
                      if isfile(join(database_path, folder, file))]

        for file in file_names:
            excel = Dispatch("Excel.Application")
            excel.Visible = True
            excel.Workbooks.Open(os.path.join(database_path, folder, file))
            time.sleep(3) # suspend 3 secs for excel to catch up python
            excel.Range("A1:XFD1048576").Select()
            excel.Selection.Copy()
            excel.Selection.PasteSpecial(Paste=-4163)
            excel.ActiveWorkbook.Save()
            excel.ActiveWorkbook.Close()


def request_fs_period(year, quarter, segment,
                      fs_type, exchange='all') -> pd.DataFrame:

    """
    This function extracts data from Github server, clean up
    and make it ready for use

    :param year: reported year
    :param quarter: reported quarter
    :param segment: allow values in request_segment_all()
    :param fs_type: allow values in request_fstype()
    :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
    :type year: int
    :type quarter: int
    :type segment: str
    :type fs_type: str
    :type exchange: str

    :return: pandas.DataFrame
    :raise ValueError: this function yet supported cashflow for
    securities companies
    """

    global database_path
    segments = request_segment_all()
    fs_types = request_fstype()

    if segment not in segments:
        raise ValueError(f'sector must be in {segments}')

    if fs_type not in fs_types:
        raise ValueError(f'sector must be in {fs_types}')

    folder = 'fs_' + segment + '_industry'
    file = fs_type + '_' + str(year) + 'q' + str(quarter) + '.xlsm'

    # create Workbook object, select active Worksheet
    raw_fiinpro \
        = openpyxl.load_workbook(
        os.path.join(database_path, folder, file)).active

    # delete StoxPlux Sign
    raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row-21,
                            amount=1000)

    # delete header rows
    raw_fiinpro.delete_rows(idx=0, amount=7)
    raw_fiinpro.delete_rows(idx=2, amount=1)

    # import to DataFrame, no column labels, no index
    clean_data = pd.DataFrame(raw_fiinpro.values)

    # assign column labels and index
    clean_data.columns = clean_data.iloc[0,:]
    clean_data.drop(index=[0], inplace=True)
    clean_data.index \
        = pd.MultiIndex.from_arrays([[year]*len(clean_data),
                                     [quarter]*len(clean_data),
                                     clean_data['Ticker'].tolist()])
    clean_data.index.set_names(['year', 'quarter', 'ticker'], inplace=True)

    # rename 2 columns
    clean_data.rename(columns=
                      {'Name':'full_name', 'Exchange':'exchange'},
                      inplace=True)

    # drop unwanted columns and index
    clean_data.drop(columns=['No', 'Ticker'], inplace=True)

    # fill na with 0s
    clean_data.fillna(0, inplace=True)

    # remove OTC
    clean_data = clean_data.loc[clean_data['exchange'] != 'OTC']


    if segment == 'bank':
        if fs_type == 'bs':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('A.','B.','C.','D.','E.','F.','G.')):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 :clean_data.columns[col].split()[0]},
                        inplace=True)
                    header.append(clean_data.columns[col])
                else:
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 :header[-1]
                                  + clean_data.columns[col].split()[0]},
                        inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += '_'
            clean_data.columns = col_list

            subheader = list('I')
            for col in range(2, len(clean_data.columns)):
                l = clean_data.columns[col].split('.')
                a = l[1]
                if a in ['I','II','III','IV','V','VI','VII']:
                    subheader.append(a)
                else:
                    name_new = l
                    name_new.insert(1,subheader[-1])
                    name_new = '.'.join(name_new)
                    clean_data.rename(
                        columns={clean_data.columns[col]:name_new},
                        inplace=True)

            clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('_'), inplace=True)

        elif fs_type == 'is':
            for col in range(2, len(clean_data.columns)):
                clean_data.rename(
                    columns={clean_data.columns[col]
                             :clean_data.columns[col].split()[0]},
                    inplace=True)

        elif fs_type == 'cfi':
            header = list()
            for col in range(len(clean_data.columns)-1,1,-1):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                      + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)

        elif fs_type == 'cfd':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : header[-1]
                                  + clean_data.columns[col].split()[0]},
                        inplace=True)
            col_list = clean_data.columns.tolist()
            duplicated = clean_data.columns.duplicated()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += 'b.'
            clean_data.columns = col_list


    elif segment == 'gen':
        # remove financial
        fin = set(request_financial_ticker()) \
              & set(clean_data.index.get_level_values(2))
        clean_data.drop(index=fin, level=2, inplace=True)
        if fs_type == 'bs':
            header = list()
            for col in range(len(clean_data.columns)-1,1,-1):
                if clean_data.columns[col]\
                    .startswith(('A.','B.','C.','D.','E.','F.','G.')):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 :clean_data.columns[col].split()[0]},
                        inplace=True)
                    header.append(clean_data.columns[col])
                else:
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 :header[-1]
                                  + clean_data.columns[col].split()[0]},
                        inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += '_'
            clean_data.columns = col_list

            subheader = list('I')
            for col in range(2, len(clean_data.columns)):
                l = clean_data.columns[col].split('.')
                a = l[1]
                if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    subheader.append(a)
                else:
                    name_new = l
                    name_new.insert(1, subheader[-1])
                    name_new = '.'.join(name_new)
                    clean_data.rename(
                        columns={clean_data.columns[col]: name_new},
                        inplace=True)

            clean_data.rename(axis=1, mapper=
            lambda x: x.rstrip('_'), inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i] and col_list[i].split('.',1)[1] \
                        not in ['I.','II.','III.','IV','V.','VI.','VII.']:
                    col_list[i] += 'b.'
            clean_data.columns = col_list

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                l = col_list[i].split('.')
                if duplicated[i] and l[1]\
                    in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    col_list[i] = l[0] + '.'
            clean_data.columns = col_list

        elif fs_type == 'is':
            for col in range(2, len(clean_data.columns)):
                clean_data.rename(
                    columns={clean_data.columns[col]
                             :clean_data.columns[col].split()[0]},
                    inplace=True)

        elif fs_type == 'cfi':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                      + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)

        elif fs_type == 'cfd':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                      + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)


    elif segment == 'ins':
        if fs_type == 'bs':
            header = list()
            for col in range(len(clean_data.columns)-1,1,-1):
                if clean_data.columns[col]\
                    .startswith(('A.','B.','C.','D.','E.','F.','G.')):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 :clean_data.columns[col].split()[0]},
                        inplace=True)
                    header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     :header[-1]
                                      + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += '_'
            clean_data.columns = col_list

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += '_'
            clean_data.columns = col_list

            subheader = list('I')
            for col in range(2, len(clean_data.columns)):
                l = clean_data.columns[col].split('.')
                a = l[1]
                if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    subheader.append(a)
                else:
                    name_new = l
                    name_new.insert(1, subheader[-1])
                    name_new = '.'.join(name_new)
                    clean_data.rename(
                        columns={clean_data.columns[col]: name_new},
                        inplace=True)

            clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('__').rstrip('_'), inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                l = col_list[i].split('.')
                if duplicated[i] and l[2] != '':
                    col_list[i] += 'b.'
            clean_data.columns = col_list

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                l = col_list[i].split('.')
                if duplicated[i] and l[1]\
                    in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    col_list[i] = l[0] + '.'
            clean_data.columns = col_list

            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                a = col_list[i].split('.')[1]
                if col_list[i].startswith(('1','2','3','4','5')):
                    col_list[i] = col_list[i].replace(a + '.', '')
            clean_data.columns = col_list

        elif fs_type == 'is':
            for col in range(2, len(clean_data.columns)):
                clean_data.rename(
                    columns={clean_data.columns[col]
                             :clean_data.columns[col].split()[0]},
                    inplace=True)
            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += 'b.'
                elif col_list[i] == '2201.':
                    col_list[i] = '20.1.'
            clean_data.columns = col_list

        elif fs_type == 'cfi':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : header[-1]
                                  + clean_data.columns[col].split()[0]},
                        inplace=True)

            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                a = col_list[i].split('.')
                if len(a[-1]) >= 5:
                    col_list[i] = '.'.join(a[:2]) + '.'
            clean_data.columns = col_list

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += 'b.'
                    break
            clean_data.columns = col_list

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += 'c.'
            clean_data.columns = col_list

        elif fs_type == 'cfd':
            header = list()
            for col in range(2, len(clean_data.columns)):
                if clean_data.columns[col]\
                    .startswith(('I','II','III','IV','V','VI','VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                      + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                a = col_list[i].split('.')
                if len(a[-1]) >= 1:
                    col_list[i] = '.'.join(a[:-1]) + '.'
            clean_data.columns = col_list


    elif segment == 'sec':
        if fs_type == 'bs':
            header = list()
            for col in range(len(clean_data.columns)-1,1,-1):
                if clean_data.columns[col] \
                        .startswith(
                    ('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.')):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : clean_data.columns[col].split()[0]},
                        inplace=True)
                    header.append(clean_data.columns[col])
                else:
                    try:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     :header[-1]
                                       + clean_data.columns[col].split()[0]},
                            inplace=True)
                    except IndexError:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     :clean_data.columns[col].split()[0]},
                            inplace=True)


            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                if duplicated[i]:
                    col_list[i] += '_'
            clean_data.columns = col_list

            subheader = list('I')
            for col in range(2, len(clean_data.columns)):
                l = clean_data.columns[col].split('.')
                a = l[1]
                if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    subheader.append(a)
                else:
                    name_new = l
                    name_new.insert(1, subheader[-1])
                    name_new = '.'.join(name_new)
                    clean_data.rename(
                        columns={clean_data.columns[col]: name_new},
                        inplace=True)

            clean_data.rename(axis=1, mapper=
            lambda x: x.rstrip('_'), inplace=True)

            duplicated = clean_data.columns.duplicated()
            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                l = col_list[i].split('.')
                if duplicated[i] and l[1]\
                    in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    col_list[i] = l[0] + '.'
            clean_data.columns = col_list

            col_list = clean_data.columns.tolist()
            for i in range(2, len(clean_data.columns)):
                l = col_list[i].split('.')
                if l[0] in ['1','2','3','4','5'] and l[1]\
                    in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                    col_list[i] = col_list[i].replace('.'+l[1], '')
            clean_data.columns = col_list

        elif fs_type == 'is':
            for col in range(2, len(clean_data.columns)):
                clean_data.rename(
                    columns={clean_data.columns[col]
                             :clean_data.columns[col].split()[0]},
                    inplace=True)

        elif fs_type == 'cfi':
            pass

        elif fs_type == 'cfd':
            pass


    clean_data.columns \
        = pd.MultiIndex.from_product([[fs_type],
                                      clean_data.columns.tolist()],
                                     names = ['fs_type', 'item'])
    if exchange != 'all':
        clean_data \
            = clean_data.loc[clean_data.loc[:,(fs_type,'exchange')]==exchange]

    print('Extracting...')
    return clean_data


def request_fs_all(segment=str) -> pd.DataFrame:

    """
    This function returns all financial statements
    of all companies in all periods

    :param segment: allow values in request_segment_all()
    :return: pandas.DataFrame
    """

    global database_path
    fs_types = request_fstype()
    periods = request_period()

    frames = list()
    for period in periods:
        for fs_type in fs_types:
            try:
                frames.append(
                    request_fs_period(int(period[:4]),
                                      int(period[-1]),
                                      segment, fs_type))
            except FileNotFoundError:
                continue

    df = pd.concat(frames, axis=1, join='outer')
    cols = df.columns.get_level_values(0)+'__'+df.columns.get_level_values(1)
    df.columns = cols
    df = df.groupby(by=cols, axis=1,
                    dropna=False, sort=False).sum(min_count=1)
    lvl_0 = [col.split('__')[0] for col in df.columns]
    lvl_1 = [col.split('__')[1] for col in df.columns]
    df.columns = pd.MultiIndex.from_arrays([lvl_0, lvl_1],
                                           names=['fs_type', 'item'])
    df.drop(columns=['exchange', 'full_name'], level=1, inplace=True)

    return df


def request_fs_ticker(ticker=str) -> pd.DataFrame:

    """
    This functions returns all financial statements
    of given ticker in all periods

    :param ticker: allow values in request_ticker_all()
    :return: pandas.DataFrame
    """

    global database_path
    segment = request_segment(ticker)
    folder = 'fs_' + segment + '_industry'
    files = listdir(join(database_path, folder))

    file_names = []
    for file in files:
        if isfile(join(database_path, folder, file)) \
                and not file.startswith('~$'):
            file_names.append(file)
    file_names = list(set(file_names))
    file_names.sort()

    refs = [(int(name[-11:-7]),int(name[-6]),name[:2]
            if name[2]=='_' else name[:3]) for name in file_names]

    fs_types = request_fstype()
    inds = list()
    for fs_type in fs_types:
        try:
            inds += request_fs_period(refs[-1][0],
                                      refs[-1][1],
                                      segment,
                                      fs_type=fs_type) \
                .xs(ticker, axis=0, level=2) \
                .drop(['full_name', 'exchange'], level=1, axis=1) \
                .columns.tolist()
        except KeyError:
            continue

    dict_ind = dict()
    for fs_type in fs_types:
        dict_ind[fs_type] = [x[1] for x in inds if x[0] == fs_type]

    fs = pd.concat(
        [
            request_fs_period(ref[0], ref[1], segment, ref[2])\
                .xs(ticker, axis=0, level=2)\
                .drop(['full_name', 'exchange'], level=1, axis=1).T\
                .set_index(pd.MultiIndex.from_product(
                    [[segment], [ref[2]], dict_ind[ref[2]]]))
            for ref in refs
        ]
    )
    fs = fs.groupby(fs.index, sort=False).sum()

    print('Finished!')
    return fs


def request_ticker(segment=str, exchange=str) -> list:

    """
    This function returns all tickers of given segment or exchange

    :param segment: allow values in request_segment_all()
    :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM']
    :return: list
    """

    if segment == 'gen':
        last_period = request_period()[-1]
        ticker_list \
            = request_fs_period(int(last_period[:4]),
                                int(last_period[-1]),
                                segment, 'is')\
            .index.get_level_values(level=2).tolist()
    else:
        fin_dict = request_financial_ticker(True)
        ticker_list = fin_dict[segment]

    return ticker_list


def request_ticker_all() -> list:

    """
    This functions returns all tickers of all segments

    :param: None
    :return: list
    """

    segments = request_segment_all()
    tickers = []
    for segment in segments:
        tickers += request_ticker(segment)

    return tickers


def request_crash(benchmark=-0.5, segment=str, period=str) -> list:

    """
    This function returns all tickers whose stock return lower than 'benchmark'
    in a given period

    :param benchmark: negative number in [-1,0]
    :param segment: allow values in request_segment_all()
    :param period: allow values in request_period()
    :return: list
    """

    returns = request_return(segment)
    crash = list()
    for ticker in returns.index:
        if returns.loc[ticker, period] <= benchmark:
            crash.append(ticker)
        else:
            continue
    return crash


def request_fs_variable(segment=str, fs_type=str) -> list:

    """
    This function returns all variables
    of given financial statement of given segment

    :param segment: allow values in request_segment_all()
    :param fs_type: allow valies in request_fstype()
    :return: list
    """

    last_period = request_period()[-1]
    variable_names \
        = request_fs_period(int(last_period[:4]),
                            int(last_period[-1]),
                            segment,
                            fs_type)\
        .columns.get_level_values(level=1).tolist()
    try:
        variable_names.remove('full_name')
        variable_names.remove('exchange')
    except ValueError:
        pass

    return variable_names


def request_period() -> list:

    """
    This function returns all periods

    :param: None
    :return: list
    """

    segments = request_segment_all()
    folders = ['fs_' + segment + '_industry' for segment in segments]

    periods = []
    for folder in folders:
        periods \
            = list(set(
            [name[-11:-5] for name in listdir(join(database_path, folder))
             if isfile(join(database_path, folder, name))]))
        periods.sort()

    return periods


def request_industry(standard=str) -> pd.DataFrame:

    """
    This funtions returns industry classification instructed by
     a given standard of all stock

    :param standard: allow values in request_industry_standard()
    :return: pandas.DataFrame
    """

    global database_path
    standards = request_industry_standard()
    st_dict = dict()
    folder = 'industry_classification'

    for st in standards:
        # create Workbook object, select active Worksheet
        raw_bloomberg \
            = openpyxl.load_workbook(
            os.path.join(database_path, folder, st+'.xlsx')
        ).active
        # delete Bloomberg Sign
        raw_bloomberg.delete_rows(idx=raw_bloomberg.max_row)
        # delete headers
        raw_bloomberg.delete_rows(idx=0, amount=2)
        clean_data = pd.DataFrame(raw_bloomberg.values)
        clean_data.iloc[1] = clean_data.iloc[0]
        clean_data.drop(index=0, inplace=True)
        # set index and columns
        clean_data.index = clean_data.iloc[:,0]
        clean_data.index.rename('ticker')
        clean_data.columns = clean_data.iloc[0,:]
        clean_data.columns.rename('level')
        # remore unwanted columns, rows
        clean_data.drop(columns=['Ticker', 'Name'],
                        index=['Ticker'], inplace=True)
        # rename columns
        clean_data.columns\
            = pd.Index(
            data = [
                clean_data.columns[i].split()[0].lower()
                    .split(' ', maxsplit=1)[0] + '_l' + str(i+1)
                for i in range(clean_data.shape[1])
            ]
        )
        # rename index
        clean_data.index \
            = pd.Index(
            data=[clean_data.index[i].split()[0]
                for i in range(clean_data.shape[0])]
        )

        st_dict[st] = clean_data

    return st_dict[standard]


def request_industry_standard() -> list:

    """
    This function returns all industry classification standards

    :param: None
    :return: list
    """
    global database_path
    folder = 'industry_classification'
    standards \
        = list(dict.fromkeys(
        [name[:-5] for name in listdir(join(database_path, folder))
         if isfile(join(database_path, folder, name))]))
    return standards


def request_industry_level(standard=str) -> list:

    """
    This function returns all levels of given industry classification standard

    :param standard: allow values in request_industry_standard()
    :return: list
    """

    levels = request_industry(standard).columns.tolist()
    return levels


def request_industry_list(standard=str, level=int) -> list:

    """
    This function returns all industry names of
    given level of given classification standard

    :param standard: allow values in request_industry_standard()
    :param level: allow values in request_industry_level() (number only)
    """

    industries \
        = request_industry(standard)[standard + '_l' + str(level)]\
        .drop_duplicates().tolist()
    return industries


def request_return(segment=str) -> pd.DataFrame:

    """
    This function returns stock returns of all tickers of given segment
    in all periods

    :param segment: None
    :return: pandas.DataFrame
    """

    periods = request_period()
    tickers = request_ticker(segment)

    returns = pd.DataFrame(data=np.zeros((len(tickers),len(periods))),
                           columns=[periods[i]
                                    for i in range(len(periods))],
                           index=[tickers[j]
                                  for j in range(len(tickers))])

    date_input = {'q1':['01-01','03-31'], 'q2':['04-01','06-30'],
                  'q3':['07-01','09-30'], 'q4':['10-01','12-28']}
    # Note: because of the 'Bday' function structure,
    #       December must be adjusted to 28 instead of 31

    price_data = dict()
    for ticker in tickers:
        try:
            price_data[ticker] = request_trading_hist(ticker)
        except (KeyError, IndexError):
            continue

    def Bday(date=str):
        date_ = datetime(year=int(date.split('-')[0]),
                         month=int(date.split('-')[1]),
                         day=int(date.split('-')[2]))
        one_day = timedelta(days=1)
        while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
            date_ = date_ + one_day

        return date_.strftime(format='%Y-%m-%d')

    for period in periods:
        for ticker in tickers:
            try:
                fromdate = Bday(period[:4] + '-' + date_input[period[-2:]][0])
                todate = Bday(period[:4] + '-' + date_input[period[-2:]][1])
            except (KeyError, IndexError):
                continue
            try:
                f = price_data[ticker]['trading_date'] == fromdate
                t = price_data[ticker]['trading_date'] == todate
                try:
                    open_price = price_data[ticker].loc[f]['close'].iloc[0]
                    close_price = price_data[ticker].loc[t]['close'].iloc[0]
                    returns.loc[ticker, period] = close_price / open_price - 1
                except (IndexError, KeyError, RuntimeWarning):
                    returns.loc[ticker, period] = np.nan
            except (IndexError, KeyError, RuntimeWarning):
                returns.loc[ticker, period] = np.nan

    returns.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)

    return returns


def request_price() -> pd.DataFrame:

    """
    This function returns stock price of all tickers in all periods

    :param: None
    :return: pandas.DataFrame
    """

    periods = request_period()
    tickers = request_ticker_all()

    prices = pd.DataFrame(data=np.zeros((len(tickers),len(periods))),
                          columns=[periods[i]
                                   for i in range(len(periods))],
                          index=[tickers[j]
                                 for j in range(len(tickers))])

    date_input = {'q1':'03-31', 'q2':'06-30',
                  'q3':'09-30', 'q4':'12-28'}
    # Note: because of the 'Bday' function structure,
    #       December must be adjusted to 28 instead of 31

    price_data = dict()
    for ticker in tickers:
        try:
            price_data[ticker] = request_trading_hist(ticker)
        except (KeyError, IndexError):
            continue

    def Bday(date=str):
        date_ = datetime(year=int(date.split('-')[0]),
                         month=int(date.split('-')[1]),
                         day=int(date.split('-')[2]))
        one_day = timedelta(days=1)
        while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
            date_ = date_ + one_day

        return date_.strftime(format='%Y-%m-%d')

    for period in periods:
        for ticker in tickers:
            try:
                date = Bday(period[:4] + '-' + date_input[period[-2:]])
            except (KeyError, IndexError):
                continue
            try:
                t = price_data[ticker]['trading_date'] == date
                try:
                    prices.loc[ticker, period] \
                        = price_data[ticker].loc[t]['close'].iloc[0]
                except (IndexError, KeyError, RuntimeWarning):
                    prices.loc[ticker, period] = np.nan
            except (IndexError, KeyError, RuntimeWarning):
                prices.loc[ticker, period] = np.nan

    prices.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)
    return prices


def ownership_structure() -> pd.DataFrame:

    """
    This function returns ownership structure of all tickers

    :param: None
    :return: pandas.DataFrame
    """

    global database_path
    folder = 'ownership'
    file = [f for f in listdir(join(database_path, folder))
            if isfile(join(database_path, folder, f))][-1]

    excel = Dispatch("Excel.Application")
    excel.Visible = True
    for wb in [wb for wb in excel.Workbooks]:
        wb.Close(True)
    excel.Workbooks.Open(os.path.join(database_path, file))
    time.sleep(3) # suspend 3 secs for excel to catch up python
    excel.Range("A1:XFD1048576").Select()
    excel.Selection.Copy()
    excel.Selection.PasteSpecial(Paste=-4163)
    excel.ActiveWorkbook.Save()
    excel.ActiveWorkbook.Close()

    # create Workbook object, select active Worksheet
    raw_fiinpro = openpyxl.load_workbook(
        os.path.join(database_path, folder, file)
    ).active
    try:
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row-21, amount=1000)
    except IndexError:
        excel = Dispatch("Excel.Application")
        excel.Visible = True
        for wb in [wb for wb in excel.Workbooks]:
            wb.Close(True)
        excel.Workbooks.Open(os.path.join(database_path, file))
        time.sleep(3) # suspend 3 secs for excel to catch up python
        excel.Range("A1:XFD1048576").Select()
        excel.Selection.Copy()
        excel.Selection.PasteSpecial(Paste=-4163)
        excel.ActiveWorkbook.Save()
        excel.ActiveWorkbook.Close()
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21,
                                amount=1000)

    # get result info
    date_of_extract = raw_fiinpro['B6'].value

    # delete header rows
    raw_fiinpro.delete_rows(idx=0, amount=7)
    raw_fiinpro.delete_rows(idx=2, amount=1)

    # import to DataFrame
    clean_data = pd.DataFrame(raw_fiinpro.values)

    # drop unwanted index and columns
    clean_data.drop(columns=[0], inplace=True)
    clean_data.drop(index=[0], inplace=True)

    # set ticker as index
    clean_data.index = clean_data.iloc[:,0]
    clean_data.drop(clean_data.columns[0], axis=1, inplace=True)

    # rename columns
    columns = ['full_name', 'exchange', 'state_share',
               'state_percent', 'frgn_share', 'frgn_percent',
               'other_share', 'other_percent', 'frgn_maxpercent',
               'frgn_maxshare','frgn_remainshare']
    clean_data.columns = columns

    return clean_data, date_of_extract


def request_trading_hist(ticker=str, fromdate=None, todate=None) \
        -> pd.DataFrame:

    """
    This function returns historical trading data of given ticker

    :param ticker: allow values in request_ticker_all()
    :param fromdate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
    :param todate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
    """

    address = 'https://api.phs.vn/market/utilities.svc/GetShareIntraday'
    pd.options.mode.chained_assignment = None
    if fromdate is not None and todate is not None:
        if datetime.strptime(fromdate, '%Y-%m-%d') < \
                datetime(year=2015, month=1, day=1):
            raise Exception('Only data since 2015/01/01 is reliable')
        else:
            try:
                r = requests.post(address,
                    data=json.dumps(
                        {'symbol': ticker,
                         'fromdate': fromdate,
                         'todate': todate}
                    ),
                    headers={'content-type': 'application/json'})
                history = pd.DataFrame(json.loads(r.json()['d']))
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
    else:
        try:
            r = requests.post(address,
                              data=json.dumps(
                                  {'symbol': ticker,
                                   'fromdate': datetime(year=2015,
                                                        month=1,
                                                        day=1) \
                                       .strftime("%Y-%m-%d"),
                                   'todate': datetime.now()
                                       .strftime("%Y-%m-%d")}
                              ),
                              headers={'content-type': 'application/json'})
            history = pd.DataFrame(json.loads(r.json()['d']))
        except KeyError:
            try:
                r = requests.post(address,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': (datetime.now()
                                                    - timedelta(days=1000)) \
                                           .strftime("%Y-%m-%d"),
                                       'todate': datetime.now()
                                           .strftime("%Y-%m-%d")}
                                  ),
                                  headers={'content-type': 'application/json'})
                history = pd.DataFrame(json.loads(r.json()['d']))
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

    history.rename(columns={'symbol':'ticker',
                            'prior_price':'ref',
                            'close_price':'close'}, inplace=True)

    def addzero(int_=str):
        if len(int_) == 1:
            int_ = '0' + int_
        else:
            pass
        return int_

    for i in range(history.shape[0]):
        history['trading_date'].iloc[i] \
            = history['trading_date'].iloc[i][:-12].split('/')[2] + "-"\
              + addzero(history['trading_date'].iloc[i][:-12] \
                        .split('/')[0]) + "-" \
              + addzero(history['trading_date'].iloc[i][:-12] \
                        .split('/')[1])

    history['ref'].iloc[0] = history['close'].iloc[0] # wait for 'open' from IT

    for col in ['ref', 'close', 'high', 'low']: # wait 'open' from IT
        for i in range(1, history.shape[0]):
            if history[col].iloc[i] == 0:
                history[col].iloc[i] = history[col].iloc[i-1]

    return history


def request_trading_intra(ticker=str, fromdate=None, todate=None) \
        -> pd.DataFrame:

    """
    This function returns intraday trading data of given ticker

    :param ticker: allow values in request_ticker_all()
    :param fromdate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
    :param todate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'

    :raise Exception: Can't extract more than 60 days
    """

    address = 'https://api.phs.vn/market/Utilities.svc/GetRealShareIntraday'
    pd.options.mode.chained_assignment = None
    if fromdate is not None and todate is not None:
        if datetime.strptime(todate, '%Y-%m-%d') \
                - datetime.strptime(fromdate, '%Y-%m-%d') > timedelta(days=60):
            raise Exception('Can\'t extract more than 60 days')
        else:
            try:
                r = requests.post(address,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': fromdate,
                                       'todate': todate}
                                  ),
                                  headers={'content-type': 'application/json'})
                #history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
                intraday = pd.DataFrame(json.loads(r.json()['d'])['intradays'])
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
    else:
        try:
            fromdate = (datetime.now()
                        - timedelta(days=60)).strftime("%Y-%m-%d")
            todate = (datetime.now()
                      - timedelta(days=0)).strftime("%Y-%m-%d")
            r = requests.post(address,
                              data=json.dumps(
                                  {'symbol': ticker,
                                   'fromdate': fromdate,
                                   'todate': todate}
                              ),
                              headers={'content-type': 'application/json'})
            #history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
            intraday = pd.DataFrame(json.loads(r.json()['d'])['intradays'])
        except KeyError:
            raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

    def datemod(date=str):
        def addzero(int_=str):
            if len(int_) == 1:
                int_ = '0' + int_
            else:
                pass
            return int_
        day = addzero(date.split('/')[0])
        month = addzero(date.split('/')[1])
        date = date.split('/')[2][:4] + '-' \
               + month + '-' \
               + day + ' ' \
               + date.split(" ",maxsplit=1)[1]
        return date

    for i in range(intraday.shape[0]):
        intraday['trading_time'].iloc[i] \
            = datemod(intraday['trading_time'].iloc[i])

    return intraday


def request_latest_close_price(ticker=str) -> float:

    """
    This function returns the latest close price of given ticker

    :param ticker: allow values in request_ticker_all()
    :return: float
    """

    close_price = request_trading_hist(ticker)['close'].iloc[-1]
    return close_price

