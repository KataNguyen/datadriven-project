import numpy as np
import pandas as pd
import openpyxl
import os
import sys
from os import listdir
from os.path import isfile, isdir, join
from win32com.client import Dispatch
import time
from datetime import datetime, timedelta
import requests
import json
import holidays
from typing import Union


###############################################################################


class fa:

    database_path \
        = join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))),
               'database')
    financials = ['bank', 'sec', 'ins']

    # Constructor:
    def __init__(self):

        folders = [f for f in listdir(self.database_path)
                   if f.startswith('fs_')]
        fs_types = []
        for folder in folders:
            fs_types = [name.split('_')[0]
                        for name in listdir(join(self.database_path, folder))
                        if isfile(join(self.database_path, folder, name))]
        fs_types = list(set(fs_types))
        self.fs_types = [x for x in fs_types if not x.startswith('~$')]
        self.fs_types.sort() # Returns all valid financial statements

        self.segments = [x.split('_')[1] for x in folders]
        self.segments.sort() # Returns all valid segments

        periods = []
        for folder in folders:
            periods \
                = list(set(
                [name[-11:-5] for name in listdir(join(self.database_path,
                                                       folder))
                 if isfile(join(self.database_path, folder, name))]))
            periods.sort()
        self.periods = periods # Returns all periods
        self.latest_period = periods[-1]

        standards \
            = list(dict.fromkeys(
            [name[:-5] for name in listdir(join(self.database_path,'industry'))
             if isfile(join(self.database_path, 'industry', name))]))
        self.standards = standards # Returns all industry classification standards


    def reload(self) -> None:

        """
        This method handles cached data in newly-added files

        :param: None
        :return: None
        """

        folder_names = [folder
                        for folder in listdir(self.database_path)
                        if isdir(join(self.database_path, folder))]

        for folder in folder_names:
            file_names = [file
                          for file in listdir(join(self.database_path, folder))
                          if isfile(join(self.database_path, folder, file))]

            for file in file_names:
                excel = Dispatch("Excel.Application")
                excel.Visible = True
                excel.Workbooks.Open(os.path.join(self.database_path,
                                                  folder, file))
                time.sleep(3)  # suspend 3 secs for excel to catch up python
                excel.Range("A1:XFD1048576").Select()
                excel.Selection.Copy()
                excel.Selection.PasteSpecial(Paste=-4163)
                excel.ActiveWorkbook.Save()
                excel.ActiveWorkbook.Close()


    def fin_tickers(self, sector_break:bool=False) \
            -> Union[list, dict]:

        """
        This method returns all tickers of financial segments

        :param sector_break: False: ignore sectors, True: show sectors
        :type sector_break: bool
        :return: list (sector_break=False), dictionary (sector_break=True)
        """

        tickers = []
        tickers_ = dict()
        for segment in self.financials:
            folder = 'fs_' + segment + '_industry'
            file = 'is_' + self.latest_period[:4] + 'q' \
                   + self.latest_period[-1] + '.xlsm'
            raw_fiinpro \
                = openpyxl.load_workbook(
                os.path.join(self.database_path, folder, file)).active
            # delete StoxPlux Sign
            raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21,
                                    amount=1000)
            # delete headers
            raw_fiinpro.delete_rows(idx=0, amount=7)
            raw_fiinpro.delete_rows(idx=2, amount=1)
            # import
            clean_data = pd.DataFrame(raw_fiinpro.values)
            clean_data.drop(index=[0], inplace=True)
            # remove OTC
            a = clean_data.loc[:, 3] != 'OTC'
            if sector_break is False:
                tickers += clean_data.loc[:, 1][a].tolist()
            else:
                tickers_[segment] = clean_data.loc[:, 1][a].tolist()
                tickers = tickers_

        return tickers


    def core(self, year:int, quarter:int, fs_type:str, segment:str,
              exchange:str='all') -> pd.DataFrame:

        """
        This method extracts data from Github server, clean up
        and make it ready for use

        :param year: reported year
        :param quarter: reported quarter
        :param fs_type: allow values in request_fstype()
        :param segment: allow values in request_segment_all()
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
        :type year: int
        :type quarter: int
        :type fs_type: str
        :type segment: str
        :type exchange: str

        :return: pandas.DataFrame
        :raise ValueError: this function yet supported cashflow for
        securities companies
        """

        if segment not in self.segments:
            raise ValueError(f'sector must be in {self.segments}')

        if fs_type not in self.fs_types:
            raise ValueError(f'sector must be in {self.fs_types}')

        folder = 'fs_' + segment + '_industry'
        file = fs_type + '_' + str(year) + 'q' + str(quarter) + '.xlsm'

        # create Workbook object, select active Worksheet
        raw_fiinpro \
            = openpyxl.load_workbook(
            os.path.join(self.database_path, folder, file)).active

        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21,
                                amount=1000)

        # delete header rows
        raw_fiinpro.delete_rows(idx=0, amount=7)
        raw_fiinpro.delete_rows(idx=2, amount=1)

        # import to DataFrame, no column labels, no index
        clean_data = pd.DataFrame(raw_fiinpro.values)

        # assign column labels and index
        clean_data.columns = clean_data.iloc[0, :]
        clean_data.drop(index=[0], inplace=True)
        clean_data.index \
            = pd.MultiIndex.from_arrays([[year] * len(clean_data),
                                         [quarter] * len(clean_data),
                                         clean_data['Ticker'].tolist()])
        clean_data.index.set_names(['year', 'quarter', 'fs'], inplace=True)

        # rename 2 columns
        clean_data.rename(columns=
                          {'Name': 'full_name', 'Exchange': 'exchange'},
                          inplace=True)

        # drop unwanted columns and index
        clean_data.drop(columns=['No', 'Ticker'], inplace=True)

        # fill na with 0s
        clean_data.fillna(0, inplace=True)

        # remove OTC
        clean_data = clean_data.loc[clean_data['exchange'] != 'OTC']

        if segment == 'bank':
            if fs_type == 'bs':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                       + clean_data.columns[col].split()[0]},
                            inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += '_'
                clean_data.columns = col_list

                subheader = list('I')
                for col in range(2, len(clean_data.columns)):
                    l = clean_data.columns[col].split('.')
                    a = l[1]
                    if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        subheader.append(a)
                    else:
                        name_new = l
                        name_new.insert(1, subheader[-1])
                        name_new = '.'.join(name_new)
                        clean_data.rename(
                            columns={clean_data.columns[col]: name_new},
                            inplace=True)

                clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('_'), inplace=True)

            elif fs_type == 'is':
                for col in range(2, len(clean_data.columns)):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : clean_data.columns[col].split()[0]},
                        inplace=True)

            elif fs_type == 'cfi':
                header = list()
                for col in range(len(clean_data.columns) - 1, 1, -1):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)

            elif fs_type == 'cfd':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                       + clean_data.columns[col].split()[0]},
                            inplace=True)
                col_list = clean_data.columns.tolist()
                duplicated = clean_data.columns.duplicated()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += 'b.'
                clean_data.columns = col_list


        elif segment == 'gen':
            # remove financial
            fin = set(self.fin_tickers()) \
                  & set(clean_data.index.get_level_values(2))
            clean_data.drop(index=fin, level=2, inplace=True)
            if fs_type == 'bs':
                header = list()
                for col in range(len(clean_data.columns) - 1, 1, -1):
                    if clean_data.columns[col] \
                            .startswith(
                        ('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                       + clean_data.columns[col].split()[0]},
                            inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += '_'
                clean_data.columns = col_list

                subheader = list('I')
                for col in range(2, len(clean_data.columns)):
                    l = clean_data.columns[col].split('.')
                    a = l[1]
                    if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        subheader.append(a)
                    else:
                        name_new = l
                        name_new.insert(1, subheader[-1])
                        name_new = '.'.join(name_new)
                        clean_data.rename(
                            columns={clean_data.columns[col]: name_new},
                            inplace=True)

                clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('_'), inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i] and col_list[i].split('.', 1)[1] \
                            not in ['I.', 'II.', 'III.', 'IV', 'V.', 'VI.',
                                    'VII.']:
                        col_list[i] += 'b.'
                clean_data.columns = col_list

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    l = col_list[i].split('.')
                    if duplicated[i] and l[1] \
                            in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        col_list[i] = l[0] + '.'
                clean_data.columns = col_list

            elif fs_type == 'is':
                for col in range(2, len(clean_data.columns)):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : clean_data.columns[col].split()[0]},
                        inplace=True)

            elif fs_type == 'cfi':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)

            elif fs_type == 'cfd':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)


        elif segment == 'ins':
            if fs_type == 'bs':
                header = list()
                for col in range(len(clean_data.columns) - 1, 1, -1):
                    if clean_data.columns[col] \
                            .startswith(
                        ('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += '_'
                clean_data.columns = col_list

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += '_'
                clean_data.columns = col_list

                subheader = list('I')
                for col in range(2, len(clean_data.columns)):
                    l = clean_data.columns[col].split('.')
                    a = l[1]
                    if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        subheader.append(a)
                    else:
                        name_new = l
                        name_new.insert(1, subheader[-1])
                        name_new = '.'.join(name_new)
                        clean_data.rename(
                            columns={clean_data.columns[col]: name_new},
                            inplace=True)

                clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('__').rstrip('_'), inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    l = col_list[i].split('.')
                    if duplicated[i] and l[2] != '':
                        col_list[i] += 'b.'
                clean_data.columns = col_list

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    l = col_list[i].split('.')
                    if duplicated[i] and l[1] \
                            in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        col_list[i] = l[0] + '.'
                clean_data.columns = col_list

                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    a = col_list[i].split('.')[1]
                    if col_list[i].startswith(('1', '2', '3', '4', '5')):
                        col_list[i] = col_list[i].replace(a + '.', '')
                clean_data.columns = col_list

            elif fs_type == 'is':
                for col in range(2, len(clean_data.columns)):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : clean_data.columns[col].split()[0]},
                        inplace=True)
                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += 'b.'
                    elif col_list[i] == '2201.':
                        col_list[i] = '20.1.'
                clean_data.columns = col_list

            elif fs_type == 'cfi':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : header[-1]
                                       + clean_data.columns[col].split()[0]},
                            inplace=True)

                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    a = col_list[i].split('.')
                    if len(a[-1]) >= 5:
                        col_list[i] = '.'.join(a[:2]) + '.'
                clean_data.columns = col_list

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += 'b.'
                        break
                clean_data.columns = col_list

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += 'c.'
                clean_data.columns = col_list

            elif fs_type == 'cfd':
                header = list()
                for col in range(2, len(clean_data.columns)):
                    if clean_data.columns[col] \
                            .startswith(
                        ('I', 'II', 'III', 'IV', 'V', 'VI', 'VII')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    a = col_list[i].split('.')
                    if len(a[-1]) >= 1:
                        col_list[i] = '.'.join(a[:-1]) + '.'
                clean_data.columns = col_list


        elif segment == 'sec':
            if fs_type == 'bs':
                header = list()
                for col in range(len(clean_data.columns) - 1, 1, -1):
                    if clean_data.columns[col] \
                            .startswith(
                        ('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.')):
                        clean_data.rename(
                            columns={clean_data.columns[col]
                                     : clean_data.columns[col].split()[0]},
                            inplace=True)
                        header.append(clean_data.columns[col])
                    else:
                        try:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : header[-1]
                                           + clean_data.columns[col].split()[
                                               0]},
                                inplace=True)
                        except IndexError:
                            clean_data.rename(
                                columns={clean_data.columns[col]
                                         : clean_data.columns[col].split()[0]},
                                inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    if duplicated[i]:
                        col_list[i] += '_'
                clean_data.columns = col_list

                subheader = list('I')
                for col in range(2, len(clean_data.columns)):
                    l = clean_data.columns[col].split('.')
                    a = l[1]
                    if a in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        subheader.append(a)
                    else:
                        name_new = l
                        name_new.insert(1, subheader[-1])
                        name_new = '.'.join(name_new)
                        clean_data.rename(
                            columns={clean_data.columns[col]: name_new},
                            inplace=True)

                clean_data.rename(axis=1, mapper=
                lambda x: x.rstrip('_'), inplace=True)

                duplicated = clean_data.columns.duplicated()
                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    l = col_list[i].split('.')
                    if duplicated[i] and l[1] \
                            in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        col_list[i] = l[0] + '.'
                clean_data.columns = col_list

                col_list = clean_data.columns.tolist()
                for i in range(2, len(clean_data.columns)):
                    l = col_list[i].split('.')
                    if l[0] in ['1', '2', '3', '4', '5'] and l[1] \
                            in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII']:
                        col_list[i] = col_list[i].replace('.' + l[1], '')
                clean_data.columns = col_list

            elif fs_type == 'is':
                for col in range(2, len(clean_data.columns)):
                    clean_data.rename(
                        columns={clean_data.columns[col]
                                 : clean_data.columns[col].split()[0]},
                        inplace=True)

            elif fs_type == 'cfi':
                pass

            elif fs_type == 'cfd':
                pass

        clean_data.columns \
            = pd.MultiIndex.from_product([[fs_type],
                                          clean_data.columns.tolist()],
                                         names=['fs_type', 'item'])
        if exchange != 'all':
            clean_data \
                = clean_data.loc[clean_data[fs_type]['exchange'] == exchange]

        print('Extracting...')
        return clean_data


    def segment(self, ticker:str) -> str:

        """
        This method returns the segment of a given ticker

        :param ticker: stock's fs
        :type ticker: str
        :return: str
        """

        segment = ''
        if ticker not in self.fin_tickers():
            segment = 'gen'
        else:
            for key in self.fin_tickers(True):
                if ticker not in self.fin_tickers(True)[key]:
                    pass
                else:
                    segment = key
                    break

        return segment


    def fs(self, ticker:str) -> pd.DataFrame:

        """
        This method returns all financial statements
        of given fs in all periods

        :param ticker: stock's ticker
        :type ticker: str
        :return: pandas.DataFrame
        """

        segment = self.segment(ticker)
        folder = 'fs_' + segment + '_industry'
        file_names = listdir(join(self.database_path, folder))
        file_names = [f for f in file_names if not f.startswith('~$')]
        file_names.sort()

        refs = [(int(name[-11:-7]), int(name[-6]), name[:2]
                    if name[2] == '_' else name[:3]) for name in file_names]

        inds = list()
        for fs_type in self.fs_types:
            try:
                inds += self.core(refs[-1][0],
                                  refs[-1][1],
                                  fs_type,
                                  self.segment(ticker)) \
                    .drop(['full_name', 'exchange'], level=1, axis=1) \
                    .columns.tolist()
            except KeyError:
                continue

        dict_ind = dict()
        for fs_type in self.fs_types:
            dict_ind[fs_type] = [x[1] for x in inds if x[0] == fs_type]

        fs = pd.concat(
            [self.core(ref[0], ref[1], ref[2], segment) \
                 .xs(ticker, axis=0, level=2) \
                 .drop(['full_name', 'exchange'], level=1, axis=1).T \
                 .set_index(pd.MultiIndex.from_product(
                 [[segment], [ref[2]], dict_ind[ref[2]]]))
             for ref in refs])
        fs = fs.groupby(fs.index, sort=False).sum()

        print('Finished!')
        return fs


    def all(self, segment:str, exchange:str='all') -> pd.DataFrame:

        """
        This method returns all financial statements of all companies

        :param segment: allow values in fa.segments
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM']
        :type segment: str
        :type exchange: str
        :return: pandas.DataFrame
        """

        frames = list()
        for period in self.periods:
            for fs_type in self.fs_types:
                try:
                    frames.append(
                        self.core(int(period[:4]),
                                  int(period[-1]),
                                  fs_type,
                                  segment,
                                  exchange))
                except FileNotFoundError:
                    continue

        df = pd.concat(frames, axis=1, join='outer')
        df = df.groupby(by=df.columns, axis=1,
                        dropna=False, sort=False).sum(min_count=1)
        df.columns = pd.MultiIndex.from_tuples(df.columns,
                                               names=['fs_type', 'item'])
        df.drop(columns=['exchange', 'full_name'], level=1, inplace=True)

        return df


    def exchanges(self) -> pd.DataFrame:

        """
        This method returns stock exchanges of all tickers

        :param: None
        :return: pandas.DataFrame
        """

        table = pd.DataFrame(columns=['exchange'])
        for segment in self.segments:
            a = self.core(int(self.latest_period[:4]),
                          int(self.latest_period[-1]), segment,
                          'is', 'all')
            a = a.xs(key='exchange', axis=1, level=1)
            a = a.droplevel(level=['year', 'quarter'], axis=0)
            a.columns = ['exchange']
            table = pd.concat([table, a])

        return table


    def exchange(self, ticker:str) -> str:

        """
        This method returns stock exchange of given stock

        :param ticker: stock's ticker
        :type ticker: str

        :return: str
        """

        exchange = self.exchanges().loc[ticker].iloc[0]

        return exchange


    def tickers(self, segment:str='all', exchange:str='all'):

        """
        This method returns all tickers of given segment or exchange

        :param segment: allow values in fa.segments or 'all'
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
        :type segment: str
        :type exchange: str
        :return: list
        """

        ticker_list = []
        if segment in self.segments:
            ticker_list \
                = self.core(int(self.latest_period[:4]),
                            int(self.latest_period[-1]),
                            'is', segment, exchange) \
                .index.get_level_values(level=2).tolist()
        elif segment == 'all':
            ticker_list = []
            for segment in self.segments:
                ticker_list \
                    += self.core(int(self.latest_period[:4]),
                                 int(self.latest_period[-1]),
                                 'is', segment, exchange) \
                    .index.get_level_values(level=2).tolist()

        return ticker_list


    def items(self, segment:str, fs_type:str) -> list:

        """
        This method returns all financial items
        of given financial statement of given segment

        :param segment: allow values in request_segment_all()
        :param fs_type: allow valies in request_fstype()
        :type segment: str
        :type fs_type: str
        :return: list
        """

        items \
            = self.core(int(self.latest_period[:4]),
                        int(self.latest_period[-1]),
                        fs_type, segment) \
            .columns.get_level_values(level=1).tolist()
        try:
            items.remove('full_name')
            items.remove('exchange')
        except ValueError:
            pass

        return items


    def classification(self, standard:str) -> pd.DataFrame:

        """
        This method returns industry classification instructed by
        a given standard of all stocks

        :param standard: allow values in fa.standards
        :type standard: str
        :return: pandas.DataFrame
        """

        st_dict = dict()
        folder = 'industry'

        for st in self.standards:
            # create Workbook object, select active Worksheet
            raw_bloomberg \
                = openpyxl.load_workbook(
                os.path.join(self.database_path, folder, st + '.xlsx')).active
            # delete Bloomberg Sign
            raw_bloomberg.delete_rows(idx=raw_bloomberg.max_row)
            # delete headers
            raw_bloomberg.delete_rows(idx=0, amount=2)
            clean_data = pd.DataFrame(raw_bloomberg.values)
            clean_data.iloc[1] = clean_data.iloc[0]
            clean_data.drop(index=0, inplace=True)
            # set index and columns
            clean_data.index = clean_data.iloc[:, 0]
            clean_data.index.rename('fs')
            clean_data.columns = clean_data.iloc[0, :]
            clean_data.columns.rename('level')
            # remore unwanted columns, rows
            clean_data.drop(columns=['Ticker', 'Name'],
                            index=['Ticker'], inplace=True)
            # rename columns
            clean_data.columns \
                = pd.Index(
                data=[clean_data.columns[i].split()[0].lower()
                          .split(' ', maxsplit=1)[0] + '_l' + str(i + 1)
                      for i in range(clean_data.shape[1])])
            # rename index
            clean_data.index \
                = pd.Index(data=[clean_data.index[i].split()[0]
                                 for i in range(clean_data.shape[0])])
            st_dict[st] = clean_data

        return st_dict[standard]


    def levels(self, standard:str) -> list:

        """
        This method returns all levels of given industry classification standard

        :param standard: allow values in fa.standards
        :type standard: str
        :return: list
        """

        levels = self.classification(standard).columns.tolist()
        return levels


    def industries(self, standard:str, level:int) -> list:

        """
        This method returns all industry names of
        given level of given classification standard

        :param standard: allow values in request_industry_standard()
        :param level: allow values in request_industry_level() (number only)
        :type standard: str
        :type level: str
        :return: list
        """

        industries \
            = self.classification(standard)[standard + '_l' + str(level)] \
            .drop_duplicates().tolist()
        return industries


    def ownerships(self) -> pd.DataFrame:

        """
        This function returns ownership structure of all tickers

        :param: None
        :return: pandas.DataFrame
        """

        folder = 'ownership'
        file = [f for f in listdir(join(self.database_path, folder))
                if isfile(join(self.database_path, folder, f))][-1]

        excel = Dispatch("Excel.Application")
        excel.Visible = True
        for wb in [wb for wb in excel.Workbooks]:
            wb.Close(True)

        # create Workbook object, select active Worksheet
        raw_fiinpro = openpyxl.load_workbook(
            os.path.join(self.database_path, folder, file)).active
        # delete StoxPlux Sign
        raw_fiinpro.delete_rows(idx=raw_fiinpro.max_row - 21, amount=1000)

        # delete header rows
        raw_fiinpro.delete_rows(idx=0, amount=7)
        raw_fiinpro.delete_rows(idx=2, amount=1)

        # import to DataFrame
        clean_data = pd.DataFrame(raw_fiinpro.values)

        # drop unwanted index and columns
        clean_data.drop(columns=[0], inplace=True)
        clean_data.drop(index=[0], inplace=True)

        # set fs as index
        clean_data.index = clean_data.iloc[:, 0]
        clean_data.drop(clean_data.columns[0], axis=1, inplace=True)

        # rename columns
        columns = ['full_name', 'exchange', 'state_share',
                   'state_percent', 'frgn_share', 'frgn_percent',
                   'other_share', 'other_percent', 'frgn_maxpercent',
                   'frgn_maxshare', 'frgn_remainshare']
        clean_data.columns = columns

        return clean_data


    def markercaps(self) -> pd.DataFrame:

       """
       This function returns market capitalization of all stocks

       :param: None
       :return: pandas.DataFrame
       """

       folder = 'ownership'
       file = [f for f in listdir(join(self.database_path, folder))
               if isfile(join(self.database_path, folder, f))][-1]



fa = fa()


###############################################################################


class ta:

    address_hist \
        = 'https://api.phs.vn/market/utilities.svc/GetShareIntraday'
    address_intra \
        = 'https://api.phs.vn/market/Utilities.svc/GetRealShareIntraday'


    def __init__(self):
        pass


    def hist(self, ticker:str, fromdate=None, todate=None) \
            -> pd.DataFrame:

        """
        This method returns historical trading data of given fs

        :param ticker: allow values in fa.tickers()
        :param fromdate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
        :param todate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
        :return: pandas.DataFrame
        """

        pd.options.mode.chained_assignment = None
        if fromdate is not None and todate is not None:
            if datetime.strptime(fromdate, '%Y-%m-%d') < \
                    datetime(year=2015, month=1, day=1):
                raise Exception('Only data since 2015-01-01 is reliable')
            else:
                try:
                    r = requests.post(self.address_hist,
                                      data=json.dumps(
                                          {'symbol': ticker,
                                           'fromdate': fromdate,
                                           'todate': todate}),
                                      headers={
                                          'content-type': 'application/json'})
                    history = pd.DataFrame(json.loads(r.json()['d']))
                except KeyError:
                    raise Exception(
                        'Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
        else:
            try:
                r = requests.post(self.address_hist,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': datetime(year=2015,
                                                            month=1,
                                                            day=1) \
                                           .strftime("%Y-%m-%d"),
                                       'todate': datetime.now()
                                           .strftime("%Y-%m-%d")}),
                                  headers={'content-type': 'application/json'})
                history = pd.DataFrame(json.loads(r.json()['d']))
            except KeyError:
                try:
                    r = requests.post(self.address_hist,
                                      data=json.dumps(
                                          {'symbol': ticker,
                                           'fromdate': (datetime.now()
                                                        - timedelta(days=1000))\
                                               .strftime("%Y-%m-%d"),
                                           'todate': datetime.now()
                                               .strftime("%Y-%m-%d")}),
                                      headers={
                                          'content-type': 'application/json'})
                    history = pd.DataFrame(json.loads(r.json()['d']))
                except KeyError:
                    raise Exception(
                        'Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

        history.rename(columns={'symbol': 'fs',
                                'open_price': 'open',
                                'prior_price': 'ref',
                                'close_price': 'close'}, inplace=True)

        def addzero(int_=str):
            if len(int_) == 1:
                int_ = '0' + int_
            else:
                pass
            return int_

        for i in range(history.shape[0]):
            history['trading_date'].iloc[i] \
                = history['trading_date'].iloc[i][:-12].split('/')[2] + "-" \
                  + addzero(history['trading_date'].iloc[i][:-12] \
                            .split('/')[0]) + "-" \
                  + addzero(history['trading_date'].iloc[i][:-12] \
                            .split('/')[1])

        history['ref'].iloc[0] = history['open'].iloc[0]

        for col in ['ref', 'open', 'close', 'high', 'low']:
            for i in range(1, history.shape[0]):
                if history[col].iloc[i] == 0:
                    history[col].iloc[i] = history[col].iloc[i - 1]

        return history


    def intra(self, ticker=str, fromdate=None, todate=None) \
            -> pd.DataFrame:

        """
        This method returns intraday trading data of given fs

        :param ticker: allow values in fa.tickers()
        :param fromdate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'
        :param todate: [optional] allow any date with format: 'yyyy-mm-dd' or 'yyyy/mm/dd'

        :return: pandas.DataFrame
        :raise Exception: Can't extract more than 60 days
        """

        pd.options.mode.chained_assignment = None
        if fromdate is not None and todate is not None:
            if datetime.strptime(todate, '%Y-%m-%d') \
                    - datetime.strptime(fromdate, '%Y-%m-%d') > timedelta(
                days=60):
                raise Exception('Can\'t extract more than 60 days')
            else:
                try:
                    r = requests.post(self.address_intra,
                                      data=json.dumps(
                                          {'symbol': ticker,
                                           'fromdate': fromdate,
                                           'todate': todate}
                                      ),
                                      headers={
                                          'content-type': 'application/json'})
                    # history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
                    intraday = pd.DataFrame(
                        json.loads(r.json()['d'])['intradays'])
                except KeyError:
                    raise Exception(
                        'Date Format Required: yyyy-mm-dd, yyyy/mm/dd')
        else:
            try:
                fromdate = (datetime.now()
                            - timedelta(days=60)).strftime("%Y-%m-%d")
                todate = (datetime.now()
                          - timedelta(days=0)).strftime("%Y-%m-%d")
                r = requests.post(self.address_intra,
                                  data=json.dumps(
                                      {'symbol': ticker,
                                       'fromdate': fromdate,
                                       'todate': todate}
                                  ),
                                  headers={'content-type': 'application/json'})
                # history = pd.DataFrame(json.loads(r.json()['d'])['histories'])
                intraday = pd.DataFrame(json.loads(r.json()['d'])['intradays'])
            except KeyError:
                raise Exception('Date Format Required: yyyy-mm-dd, yyyy/mm/dd')

        def datemod(date=str):
            def addzero(int_=str):
                if len(int_) == 1:
                    int_ = '0' + int_
                else:
                    pass
                return int_

            day = addzero(date.split('/')[1])
            month = addzero(date.split('/')[0])
            date = date.split('/')[2][:4] + '-' \
                   + month + '-' \
                   + day + ' ' \
                   + date.split(" ", maxsplit=1)[1]
            return date

        for i in range(intraday.shape[0]):
            intraday['trading_time'].iloc[i] \
                = datemod(intraday['trading_time'].iloc[i])

        return intraday


    def prices(self, segment:str='all', exchange:str='all')\
            -> pd.DataFrame:

        """
        This method returns stock price of all tickers in all periods

        :param segment: allow values in fa.segments
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
        :type segment: str
        :type exchange: str
        :return: pandas.DataFrame
        """

        periods = fa.periods
        tickers = fa.tickers(segment, exchange)

        prices = pd.DataFrame(data=np.zeros((len(tickers), len(periods))),
                              columns=[periods[i]
                                       for i in range(len(periods))],
                              index=[tickers[j]
                                     for j in range(len(tickers))])

        date_input = {'q1': '03-31', 'q2': '06-30',
                      'q3': '09-30', 'q4': '12-28'}
        # Note: because of the 'Bday' function structure,
        #       December must be adjusted to 28 instead of 31

        price_data = dict()
        for ticker in tickers:
            try:
                price_data[ticker] = self.hist(ticker)
            except (KeyError, IndexError):
                continue

        def Bday(date=str):
            date_ = datetime(year=int(date.split('-')[0]),
                             month=int(date.split('-')[1]),
                             day=int(date.split('-')[2]))
            one_day = timedelta(days=1)
            while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
                date_ = date_ + one_day

            return date_.strftime(format='%Y-%m-%d')

        for period in periods:
            for ticker in tickers:
                try:
                    date = Bday(period[:4] + '-' + date_input[period[-2:]])
                except (KeyError, IndexError):
                    continue
                try:
                    t = price_data[ticker]['trading_date'] == date
                    try:
                        prices.loc[ticker, period] \
                            = price_data[ticker].loc[t]['close'].iloc[0]
                    except (IndexError, KeyError, RuntimeWarning):
                        prices.loc[ticker, period] = np.nan
                except (IndexError, KeyError, RuntimeWarning):
                    prices.loc[ticker, period] = np.nan

        prices.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)
        return prices


    def returns(self, segment:str='all', exchange:str='all') \
            -> pd.DataFrame:

        """
        This method returns stock returns of all tickers of given segment
        in all periods

        :param segment: allow values in fa.segments
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
        :type segment: str
        :type exchange: str
        :return: pandas.DataFrame
        """

        periods = fa.periods
        tickers = fa.tickers(segment, exchange)

        returns = pd.DataFrame(data=np.zeros((len(tickers),len(periods))),
                               columns=[periods[i]
                                        for i in range(len(periods))],
                               index=[tickers[j]
                                      for j in range(len(tickers))])

        date_input = {'q1':['01-01','03-31'], 'q2':['04-01','06-30'],
                      'q3':['07-01','09-30'], 'q4':['10-01','12-28']}
        # Note: because of the 'Bday' function structure,
        #       December must be adjusted to 28 instead of 31

        price_data = dict()
        for ticker in tickers:
            try:
                price_data[ticker] = self.hist(ticker)
            except (KeyError, IndexError):
                continue

        def Bday(date=str):
            date_ = datetime(year=int(date.split('-')[0]),
                             month=int(date.split('-')[1]),
                             day=int(date.split('-')[2]))
            one_day = timedelta(days=1)
            while date_.weekday() in holidays.WEEKEND or date_ in holidays.VN():
                date_ = date_ + one_day

            return date_.strftime(format='%Y-%m-%d')

        for period in periods:
            for ticker in tickers:
                try:
                    fromdate \
                        = Bday(period[:4] + '-' + date_input[period[-2:]][0])
                    todate \
                        = Bday(period[:4] + '-' + date_input[period[-2:]][1])
                except (KeyError, IndexError):
                    continue
                try:
                    f = price_data[ticker]['trading_date'] == fromdate
                    t = price_data[ticker]['trading_date'] == todate
                    try:
                        open_price \
                            = price_data[ticker].loc[f]['close'].iloc[0]
                        close_price \
                            = price_data[ticker].loc[t]['close'].iloc[0]
                        returns.loc[ticker, period]\
                            = close_price / open_price - 1
                    except (IndexError, KeyError, RuntimeWarning):
                        returns.loc[ticker, period] = np.nan
                except (IndexError, KeyError, RuntimeWarning):
                    returns.loc[ticker, period] = np.nan

        returns.replace([np.inf, -np.inf, -1, 1], np.nan, inplace=True)

        return returns


    def crash(self, benchmark=-0.5, period:str=fa.latest_period,
              segment:str='all', exchange:str='all') -> list:

        """
        This method returns all tickers whose stock return
        lower than 'benchmark' in a given period

        :param benchmark: negative number in [-1,0]
        :param period: allow values in fa.periods
        :param segment: allow values in fa.segments
        :param exchange: allow values in ['HOSE', 'HNX', 'UPCOM'] or 'all'
        :return: list
        """

        returns = self.returns(segment, exchange)
        crash = list()
        for ticker in returns.index:
            if returns.loc[ticker, period] <= benchmark:
                crash.append(ticker)
            else:
                pass

        return crash


    def now(self, ticker:str) -> float:

        """
        This method returns the latest matched price of given ticker

        :param ticker: allow values in fa.tickers()
        :return: float
        """

        oneday = timedelta(days=1)
        before = datetime.now() - timedelta(days=1)
        while datetime.now().weekday() in holidays.WEEKEND \
                or datetime.now() in holidays.VN():
            before -= oneday

        before = before.strftime("%Y-%m-%d")
        today = datetime.now().strftime("%Y-%m-%d")
        now = self.intra(ticker, before, today)['price'].iloc[-1]

        return now


    def last(self, ticker:str) -> float:

        """
        This method returns the latest close price of given ticker

        :param ticker: allow values in fa.tickers()
        :type ticker: str
        :return: float
        """

        close = self.hist(ticker)['close'].iloc[-1]
        return close


ta = ta()


###############################################################################